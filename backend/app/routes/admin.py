import io
from datetime import date, timedelta, datetime
from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models.user import User
from app.models.daily_card import DailyCard
from app.models.halqa import Halqa
from app.utils.decorators import role_required

admin_bp = Blueprint("admin", __name__)


# ─── User Management ──────────────────────────────────────────────────────────


@admin_bp.route("/registrations", methods=["GET"])
@jwt_required()
@role_required("super_admin")
def get_registrations():
    """Get all pending registrations."""
    status_filter = request.args.get("status", "pending")
    if status_filter == "all":
        users = User.query.order_by(User.created_at.desc()).all()
    else:
        users = User.query.filter_by(status=status_filter).order_by(User.created_at.desc()).all()
    return jsonify({"users": [u.to_dict() for u in users]}), 200


@admin_bp.route("/registration/<int:user_id>/approve", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def approve_registration(user_id):
    """Approve a registration request."""
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404

    user.status = "active"
    user.rejection_note = None
    db.session.commit()
    return jsonify({"message": "تم قبول الطلب", "user": user.to_dict()}), 200


@admin_bp.route("/registration/<int:user_id>/reject", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def reject_registration(user_id):
    """Reject a registration request."""
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404

    data = request.get_json() or {}
    user.status = "rejected"
    user.rejection_note = data.get("note", "")
    db.session.commit()
    return jsonify({"message": "تم رفض الطلب", "user": user.to_dict()}), 200


@admin_bp.route("/users", methods=["GET"])
@jwt_required()
@role_required("super_admin")
def get_all_users():
    """Get all users with optional filters."""
    status = request.args.get("status")
    gender = request.args.get("gender")
    halqa_id = request.args.get("halqa_id")
    search = request.args.get("search", "")

    query = User.query

    if status:
        query = query.filter_by(status=status)
    if gender:
        query = query.filter_by(gender=gender)
    if halqa_id:
        query = query.filter_by(halqa_id=int(halqa_id))
    if search:
        query = query.filter(
            db.or_(
                User.full_name.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%"),
            )
        )

    users = query.order_by(User.created_at.desc()).all()
    return jsonify({"users": [u.to_dict() for u in users]}), 200


@admin_bp.route("/user/<int:user_id>", methods=["GET"])
@jwt_required()
@role_required("super_admin")
def get_user(user_id):
    """Get user details."""
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404
    return jsonify({"user": user.to_dict()}), 200


@admin_bp.route("/user/<int:user_id>", methods=["PUT"])
@jwt_required()
@role_required("super_admin")
def update_user(user_id):
    """Update user details."""
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404

    data = request.get_json()
    allowed = ["full_name", "gender", "age", "phone", "country", "referral_source", "status", "halqa_id"]
    for field in allowed:
        if field in data:
            setattr(user, field, data[field])

    db.session.commit()
    return jsonify({"message": "تم تحديث البيانات", "user": user.to_dict()}), 200


@admin_bp.route("/user/<int:user_id>/reset-password", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def admin_reset_password(user_id):
    """Reset user password by admin."""
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404

    data = request.get_json()
    new_password = data.get("new_password", "")
    if len(new_password) < 6:
        return jsonify({"error": "كلمة المرور يجب أن تكون 6 أحرف على الأقل"}), 400

    user.set_password(new_password)
    db.session.commit()
    return jsonify({"message": "تم إعادة تعيين كلمة المرور"}), 200


@admin_bp.route("/user/<int:user_id>/withdraw", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def withdraw_user(user_id):
    """Mark user as withdrawn."""
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404

    user.status = "withdrawn"
    db.session.commit()
    return jsonify({"message": "تم سحب المشارك", "user": user.to_dict()}), 200


@admin_bp.route("/user/<int:user_id>/activate", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def activate_user(user_id):
    """Re-activate user."""
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404

    user.status = "active"
    db.session.commit()
    return jsonify({"message": "تم تفعيل المشارك", "user": user.to_dict()}), 200


# ─── Role Management ──────────────────────────────────────────────────────────


@admin_bp.route("/user/<int:user_id>/set-role", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def set_user_role(user_id):
    """Set user role (supervisor, super_admin, participant)."""
    admin_id = get_jwt_identity()
    admin = User.query.get(admin_id)
    target = User.query.get(user_id)

    if not target:
        return jsonify({"error": "المستخدم غير موجود"}), 404

    data = request.get_json()
    new_role = data.get("role")

    if new_role not in ["participant", "supervisor", "super_admin"]:
        return jsonify({"error": "الصلاحية غير صالحة"}), 400

    # Only primary admin can manage super_admin roles
    primary_email = current_app.config.get("SUPER_ADMIN_EMAIL", "").lower()
    if new_role == "super_admin" or target.role == "super_admin":
        if admin.email != primary_email:
            return jsonify({"error": "فقط المشرف الرئيسي يمكنه إدارة صلاحيات السوبر آدمن"}), 403

    target.role = new_role
    db.session.commit()
    return jsonify({"message": "تم تحديث الصلاحية", "user": target.to_dict()}), 200


# ─── Halqa Management ─────────────────────────────────────────────────────────


@admin_bp.route("/halqas", methods=["GET"])
@jwt_required()
@role_required("super_admin")
def get_halqas():
    """Get all halqas."""
    halqas = Halqa.query.all()
    return jsonify({"halqas": [h.to_dict() for h in halqas]}), 200


@admin_bp.route("/halqa", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def create_halqa():
    """Create a new halqa."""
    data = request.get_json()
    name = data.get("name", "").strip()

    if not name:
        return jsonify({"error": "اسم الحلقة مطلوب"}), 400

    if Halqa.query.filter_by(name=name).first():
        return jsonify({"error": "اسم الحلقة موجود مسبقاً"}), 400

    halqa = Halqa(name=name, supervisor_id=data.get("supervisor_id"))
    db.session.add(halqa)
    db.session.commit()
    return jsonify({"message": "تم إنشاء الحلقة", "halqa": halqa.to_dict()}), 201


@admin_bp.route("/halqa/<int:halqa_id>", methods=["PUT"])
@jwt_required()
@role_required("super_admin")
def update_halqa(halqa_id):
    """Update halqa details."""
    halqa = Halqa.query.get(halqa_id)
    if not halqa:
        return jsonify({"error": "الحلقة غير موجودة"}), 404

    data = request.get_json()
    if "name" in data:
        halqa.name = data["name"]
    if "supervisor_id" in data:
        halqa.supervisor_id = data["supervisor_id"]

    db.session.commit()
    return jsonify({"message": "تم تحديث الحلقة", "halqa": halqa.to_dict()}), 200


@admin_bp.route("/halqa/<int:halqa_id>/assign-members", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def assign_members_to_halqa(halqa_id):
    """Assign members to a halqa."""
    halqa = Halqa.query.get(halqa_id)
    if not halqa:
        return jsonify({"error": "الحلقة غير موجودة"}), 404

    data = request.get_json()
    user_ids = data.get("user_ids", [])

    for uid in user_ids:
        user = User.query.get(uid)
        if user:
            user.halqa_id = halqa_id

    db.session.commit()
    return jsonify({"message": "تم تعيين المشاركين"}), 200


@admin_bp.route("/user/<int:user_id>/assign-halqa", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def assign_user_halqa(user_id):
    """Assign a single user to a halqa."""
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404

    data = request.get_json()
    user.halqa_id = data.get("halqa_id")
    db.session.commit()
    return jsonify({"message": "تم تعيين الحلقة", "user": user.to_dict()}), 200


# ─── Analytics Dashboard ──────────────────────────────────────────────────────


@admin_bp.route("/analytics", methods=["GET"])
@jwt_required()
@role_required("super_admin")
def get_analytics():
    """Get comprehensive analytics."""
    # Filters
    gender = request.args.get("gender")
    halqa_id = request.args.get("halqa_id")
    supervisor_name = request.args.get("supervisor")
    member_name = request.args.get("member")
    min_pct = request.args.get("min_pct", type=float)
    max_pct = request.args.get("max_pct", type=float)
    period = request.args.get("period", "all")  # weekly, monthly, all
    sort_by = request.args.get("sort_by", "score")  # score, name
    sort_order = request.args.get("sort_order", "desc")

    # Base query for active users
    query = User.query.filter_by(status="active")

    if gender:
        query = query.filter_by(gender=gender)
    if halqa_id:
        query = query.filter_by(halqa_id=int(halqa_id))
    if member_name:
        query = query.filter(User.full_name.ilike(f"%{member_name}%"))
    if supervisor_name:
        halqas = Halqa.query.join(User, Halqa.supervisor_id == User.id).filter(
            User.full_name.ilike(f"%{supervisor_name}%")
        ).all()
        halqa_ids = [h.id for h in halqas]
        if halqa_ids:
            query = query.filter(User.halqa_id.in_(halqa_ids))

    users = query.all()

    # Date range
    today = date.today()
    if period == "weekly":
        start_date = today - timedelta(days=today.weekday())
    elif period == "monthly":
        start_date = today.replace(day=1)
    else:
        start_date = None

    results = []
    for u in users:
        card_query = DailyCard.query.filter_by(user_id=u.id)
        if start_date:
            card_query = card_query.filter(DailyCard.date >= start_date)

        cards = card_query.all()
        total = sum(c.total_score for c in cards)
        max_total = sum(c.max_score for c in cards) if cards else 0
        pct = round((total / max_total) * 100, 1) if max_total > 0 else 0

        # Apply percentage filter
        if min_pct is not None and pct < min_pct:
            continue
        if max_pct is not None and pct > max_pct:
            continue

        results.append({
            "user_id": u.id,
            "full_name": u.full_name,
            "gender": u.gender,
            "halqa_name": u.halqa.name if u.halqa else "بدون حلقة",
            "supervisor_name": u.halqa.supervisor.full_name if u.halqa and u.halqa.supervisor else "-",
            "total_score": total,
            "max_score": max_total,
            "percentage": pct,
            "cards_count": len(cards),
        })

    # Sorting
    if sort_by == "name":
        results.sort(key=lambda x: x["full_name"], reverse=(sort_order == "desc"))
    else:
        results.sort(key=lambda x: x["total_score"], reverse=(sort_order == "desc"))

    # Add ranks
    for i, r in enumerate(results):
        r["rank"] = i + 1

    # Summary stats
    total_active = User.query.filter_by(status="active").count()
    total_pending = User.query.filter_by(status="pending").count()
    total_halqas = Halqa.query.count()

    return jsonify({
        "results": results,
        "summary": {
            "total_active": total_active,
            "total_pending": total_pending,
            "total_halqas": total_halqas,
            "filtered_count": len(results),
        },
    }), 200


# ─── Import / Export ──────────────────────────────────────────────────────────


@admin_bp.route("/export", methods=["GET"])
@jwt_required()
@role_required("super_admin")
def export_data():
    """Export analytics data as CSV or XLSX."""
    import csv
    from openpyxl import Workbook

    format_type = request.args.get("format", "csv")

    # Get filtered data (reuse analytics logic)
    gender = request.args.get("gender")
    halqa_id = request.args.get("halqa_id")
    period = request.args.get("period", "all")

    query = User.query.filter_by(status="active")
    if gender:
        query = query.filter_by(gender=gender)
    if halqa_id:
        query = query.filter_by(halqa_id=int(halqa_id))

    users = query.all()
    today = date.today()
    if period == "weekly":
        start_date = today - timedelta(days=today.weekday())
    elif period == "monthly":
        start_date = today.replace(day=1)
    else:
        start_date = None

    rows = []
    for u in users:
        card_query = DailyCard.query.filter_by(user_id=u.id)
        if start_date:
            card_query = card_query.filter(DailyCard.date >= start_date)
        cards = card_query.all()
        total = sum(c.total_score for c in cards)
        max_total = sum(c.max_score for c in cards) if cards else 0
        pct = round((total / max_total) * 100, 1) if max_total > 0 else 0

        rows.append({
            "الاسم": u.full_name,
            "الجنس": u.gender,
            "الحلقة": u.halqa.name if u.halqa else "-",
            "المشرف": u.halqa.supervisor.full_name if u.halqa and u.halqa.supervisor else "-",
            "مجموع النقاط": total,
            "الحد الأعلى": max_total,
            "النسبة %": pct,
            "عدد البطاقات": len(cards),
        })

    rows.sort(key=lambda x: x["مجموع النقاط"], reverse=True)

    if format_type == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "النتائج"

        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row[h] for h in headers])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="ramadan_results.xlsx",
        )
    else:
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

        return output.getvalue(), 200, {
            "Content-Type": "text/csv; charset=utf-8",
            "Content-Disposition": "attachment; filename=ramadan_results.csv",
        }


@admin_bp.route("/import", methods=["POST"])
@jwt_required()
@role_required("super_admin")
def import_users():
    """Import users from Excel file."""
    from openpyxl import load_workbook

    if "file" not in request.files:
        return jsonify({"error": "لم يتم رفع ملف"}), 400

    file = request.files["file"]
    wb = load_workbook(file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    required_headers = ["الاسم", "الجنس", "العمر", "الهاتف", "البريد", "الدولة"]

    for h in required_headers:
        if h not in headers:
            return jsonify({"error": f"العمود {h} مفقود من الملف"}), 400

    imported = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        try:
            email = str(row_data.get("البريد", "")).lower().strip()
            if not email or User.query.filter_by(email=email).first():
                errors.append(f"صف {row_idx}: بريد مكرر أو فارغ")
                continue

            user = User(
                full_name=str(row_data.get("الاسم", "")).strip(),
                gender=str(row_data.get("الجنس", "")).strip(),
                age=int(row_data.get("العمر", 0)),
                phone=str(row_data.get("الهاتف", "")).strip(),
                email=email,
                country=str(row_data.get("الدولة", "")).strip(),
                referral_source=str(row_data.get("المصدر", "")).strip(),
                status="active",
                role="participant",
            )
            user.set_password("123456")  # Default password
            db.session.add(user)
            imported += 1
        except Exception as e:
            errors.append(f"صف {row_idx}: {str(e)}")

    db.session.commit()
    return jsonify({
        "message": f"تم استيراد {imported} مشارك",
        "errors": errors,
    }), 200


@admin_bp.route("/import-template", methods=["GET"])
@jwt_required()
@role_required("super_admin")
def get_import_template():
    """Download import template."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "قالب الاستيراد"
    ws.sheet_view.rightToLeft = True

    headers = ["الاسم", "الجنس", "العمر", "الهاتف", "البريد", "الدولة", "المصدر"]
    ws.append(headers)

    # Example row
    ws.append(["أحمد محمد علي", "ذكر", 25, "+966500000000", "ahmed@example.com", "السعودية", "صديق"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="import_template.xlsx",
    )
    